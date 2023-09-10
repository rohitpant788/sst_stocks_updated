import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
import helper
import config

pd.set_option('display.max_columns', None)


def load_workbook(file_path):
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {file_path}")


def calculate_metrics(df):
    df['low_20'] = df['Low'].rolling(20).min()
    df['low_3'] = df['Low'].rolling(3).min()
    df['high_20_day'] = df['High'].rolling(20).max()
    return df


def update_main_sheet(sheet, sym, metrics):
    # Update the main sheet with calculated metrics
    row = sheet.max_row + 1
    sheet.cell(row, 1).value = sym
    # Add more cell updates as needed for your specific use case

    # Calculate whether Close price is greater than 20-Day Moving Average
    close_price = metrics['Close'].iloc[-1]
    moving_average_20 = metrics['low_20'].iloc[-1]

    if close_price > moving_average_20:
        sheet.cell(row, 10).value = "Yes"
    else:
        sheet.cell(row, 10).value = "No"


def main():
    try:
        new_symbols_for_tracking = []
        wb = load_workbook(config.WORKBOOK_FOR_SST)
        main_sheet = wb[config.WORKBOOK_MAIN_SHEET]

        for row_index in range(2, main_sheet.max_row + 1):
            sym = main_sheet.cell(row_index, 1).value
            candles = helper.get_historical_data(sym, 40)
            df = pd.DataFrame(candles)
            df = calculate_metrics(df)

            # Update the main sheet with calculated metrics
            update_main_sheet(main_sheet, sym, df)

            # Check conditions for tracking new symbols
            if df['low_3'].iloc[-1] == df['low_20'].iloc[-1]:
                new_symbols_for_tracking.append(sym)

        # Save the updated workbook
        wb.save(config.WORKBOOK_FOR_SST)

        # Handle tracking sheet logic here...
        # You can add code here to update the tracking sheet with new_symbols_for_tracking

    except Exception as e:
        print(f"An error occurred: {str(e)}")


if __name__ == "__main__":
    main()
