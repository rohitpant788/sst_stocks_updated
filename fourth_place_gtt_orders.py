import openpyxl
from datetime import date
import helper
from config import WORKBOOK_FOR_SST, WORKBOOK_TRACKING_SHEET, PLACE_GTT_ORDERS

# Constants
SYMBOL_COL = 1
DIFF_20DH_CMP_COL = 4
PRICE_GRT_20_MA_COL = 9
MAX_DIFF_RANGE = 5
YES_VALUE = 'Yes'


def get_cell_value(sheet, row, col):
    return sheet.cell(row, col).value

def init():
    # This login is only required when running independently
    # helper.login_kite_in_browser()

    # Read Tracking sheet and place orders
    # Loading the workbook for SST
    wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

    # Loading the Tracking sheet from Workbook
    sh2 = wb[WORKBOOK_TRACKING_SHEET]

    # Storing the total number of row count in row variable
    row = sh2.max_row

    helper.delete_all_gtt_orders()

    # Iterating all the rows in sst.xlsx and place GTT orders.
    if PLACE_GTT_ORDERS == 'Yes':
        d = date.today()
        for i in range(2, row + 1):  # Start from 2 to skip the header row
            sym = get_cell_value(sh2, i, SYMBOL_COL)
            diff_20DH_CMP = int(get_cell_value(sh2, i, DIFF_20DH_CMP_COL))
            price_grt_20_ma = get_cell_value(sh2, i, PRICE_GRT_20_MA_COL)

            if 0 <= diff_20DH_CMP <= MAX_DIFF_RANGE and price_grt_20_ma == YES_VALUE:
                # If the GTT Order Price column in excel is None then the 20 Day High is the GTT Price
                gtt_order_price = get_cell_value(sh2, i, 3)
                # Updating the 20 DH in the GTT order price column
                sh2.cell(i, 6).value = gtt_order_price
                sh2.cell(i, 5).value = d
                # Placing a buy order and this is a fresh scenario (Iff Price > 20MA)
                helper.place_buy_order(sym, gtt_order_price)
                print(f'Buy order placed for {sym}')

    # Saving the final workbook
    wb.save(WORKBOOK_FOR_SST)


def main():
    init()

main()

# if __name__ == "__main__":
#     main()
