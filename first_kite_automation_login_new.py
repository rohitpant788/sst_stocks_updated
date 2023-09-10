import csv
import os
import time
import openpyxl
from selenium.webdriver.common.by import By
import helper
import kite_automation_config
from config import WORKBOOK_FOR_SST, WORKBOOK_INVESTMENT_SHEET, WORKBOOK_TRACKING_SHEET

# Function to download holdings and return the file path
def download_holdings(driver):
    # Click Kite Holdings BTN
    if helper.click_label_by_xpath(driver, kite_automation_config.KITE_HOLDINGS_DOWNLOAD_BTN_XPATH, 10):
        time.sleep(kite_automation_config.DELAY_IN_SEC + 5)
        file_path = r'C:\Users\Rohit\Downloads\holdings.csv'
        return file_path

    print("Clicking Kite Holdings BTN failed.")
    return None

# Function to process and update the workbook with holdings data
def update_workbook_with_holdings(file_path, wb):
    if not file_path:
        return

    sh4 = wb[WORKBOOK_INVESTMENT_SHEET]
    with open(file_path) as file:
        reader = csv.reader(file)
        i = 1
        for row in reader:
            helper.sym_where_we_have_position.add(row[0])
            for col_num, value in enumerate(row, start=1):
                sh4.cell(i, col_num).value = value
            i += 1
    os.remove(file_path)

# Function to remove rows in Tracking sheet where positions exist
def remove_existing_positions(wb):
    sh2 = wb[WORKBOOK_TRACKING_SHEET]
    max_row = sh2.max_row
    for i in range(1, max_row):
        sym = sh2.cell(i + 1, 1).value
        if sym in helper.sym_where_we_have_position:
            sh2.delete_rows(idx=i + 1)
            print(f'Position in {sym} therefore deleting it from Tracking sheet')

# Main script
def first_script():
    # Login to Kite Browser
    helper.login_kite_in_browser()
    driver = helper.get_driver()
    time.sleep(kite_automation_config.DELAY_IN_SEC)

    # Click Risk Disclosures and Holdings Menu
    helper.click_label_by_xpath(driver, kite_automation_config.KITE_RISK_DISCLOSURE_XPATH)

    #Click Holding Menu
    helper.click_label_by_xpath(driver, kite_automation_config.KITE_HOLDINGS_MENU_XPATH)

    # Download holdings and get file path
    holdings_file_path = download_holdings(driver)

    if holdings_file_path:
        # Load the workbook for SST
        wb = openpyxl.load_workbook(WORKBOOK_FOR_SST)

        # Update workbook with holdings data and remove existing positions
        update_workbook_with_holdings(holdings_file_path, wb)
        remove_existing_positions(wb)

        # Save the updated workbook and clean up
        wb.save(WORKBOOK_FOR_SST)
        driver.quit()

first_script()
# if __name__ == "__main__":
#     first_script()
